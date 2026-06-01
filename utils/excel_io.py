"""Excel 导入导出 — openpyxl"""
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    openpyxl = None

from database import get_conn


def export_herbs(filepath, where_clause="", params=None):
    """导出药材表到 Excel"""
    if openpyxl is None:
        raise ImportError("openpyxl 未安装，请执行 pip install openpyxl")

    conn = get_conn()
    cur = conn.cursor()
    sql = "SELECT name,alias,name2,supplier,purchase_note,purchase_price,sell_price,purchase_date,shelf_life_days,expiry_date,stock_qty FROM herbs"
    if where_clause:
        sql += " " + where_clause
    cur.execute(sql, params or [])
    rows = cur.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "药材数据"

    headers = ["品名","编号","品名Ⅱ","厂家","采购","价格","售价(/g)","进货日期","保存时长(天)","效期","数量(kg)"]
    header_fill = PatternFill(start_color="5C3322", end_color="5C3322", fill_type="solid")
    header_font = Font(name="Microsoft YaHei", size=11, color="D7CCC8", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D7CCC8"),
        right=Side(style="thin", color="D7CCC8"),
        top=Side(style="thin", color="D7CCC8"),
        bottom=Side(style="thin", color="D7CCC8"),
    )

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin_border
            cell.font = Font(name="Microsoft YaHei", size=10)

    ws.auto_filter.ref = ws.dimensions
    wb.save(filepath)


def import_herbs(filepath):
    """从 Excel 导入药材，返回 (success_count, errors)"""
    if openpyxl is None:
        raise ImportError("openpyxl 未安装，请执行 pip install openpyxl")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    headers = _read_headers(ws)
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    # Header name aliases -> canonical field
    FIELD_KEYS = {
        "name": ["品名", "名称"],
        "alias": ["编号", "别名"],
        "name2": ["品名Ⅱ", "品名2"],
        "supplier": ["厂家", "药商"],
        "stock_qty": ["数量", "数量(kg)", "库存", "库存(g)"],
        "purchase_note": ["采购"],
        "expiry_date": ["效期", "过期时间"],
        "purchase_price": ["价格", "进价"],
        "sell_price": ["售价(/g)", "售价/g", "售价/克", "售价"],
        "purchase_date": ["进货日期"],
        "shelf_life_days": ["保存时长(天)", "保存时长"],
    }

    def _col(field):
        for alias in FIELD_KEYS.get(field, []):
            if alias in headers:
                return headers[alias]
        return None

    def _str(row, field):
        ci = _col(field)
        if ci is None:
            return ""
        val = row[ci] if ci < len(row) else None
        if val is None:
            return ""
        if isinstance(val, float) and field in ("stock_qty", "purchase_price", "sell_price"):
            return str(val)
        return str(val).strip()

    def _float(row, field):
        ci = _col(field)
        if ci is None:
            return 0.0
        val = row[ci] if ci < len(row) else None
        try:
            return float(val) if val is not None else 0.0
        except (ValueError, TypeError):
            return 0.0

    def _int(row, field):
        ci = _col(field)
        if ci is None:
            return 365
        val = row[ci] if ci < len(row) else None
        try:
            return int(val) if val is not None else 365
        except (ValueError, TypeError):
            return 365

    def _date(row, field):
        import datetime as _dt
        ci = _col(field)
        if ci is None:
            return ""
        val = row[ci] if ci < len(row) else None
        if val is None:
            return ""
        if isinstance(val, _dt.datetime):
            return val.strftime("%Y-%m-%d")
        s = str(val).strip()
        if ' ' in s:
            s = s.split(' ')[0]
        return s

    name_col = _col("name")
    if name_col is None:
        return 0, ["无法识别表头：缺少'品名'列"]

    conn = get_conn()
    cur = conn.cursor()
    success = 0
    errors = []
    for ri, row in enumerate(rows, 2):
        if not row:
            continue
        name = _str(row, "name")
        if not name:
            errors.append(f"第{ri}行：名称为空")
            continue
        try:
            alias = _str(row, "alias")
            name2 = _str(row, "name2")
            if not name2:
                # Fallback: if no 品名Ⅱ header, try the last column beyond known headers
                max_ci = max(headers.values()) if headers else 0
                if len(row) > max_ci + 1:
                    last_val = row[max_ci + 1]
                    if last_val and str(last_val).strip():
                        name2 = str(last_val).strip()
                if not name2:
                    name2 = name
            supplier = _str(row, "supplier")
            purchase_note = _str(row, "purchase_note")
            purchase_price = _float(row, "purchase_price")
            sell_price = _float(row, "sell_price")
            purchase_date = _date(row, "purchase_date")
            shelf_life = _int(row, "shelf_life_days")
            expiry_date = _date(row, "expiry_date")
            stock_qty = _float(row, "stock_qty")

            cur.execute(
                "INSERT INTO herbs(name,alias,name2,supplier,purchase_note,purchase_price,sell_price,purchase_date,shelf_life_days,expiry_date,stock_qty) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
                [name, alias, name2, supplier, purchase_note, purchase_price, sell_price, purchase_date, shelf_life, expiry_date, stock_qty]
            )
            success += 1
        except Exception as e:
            errors.append(f"第{ri}行：{e}")
    conn.commit()
    conn.close()
    return success, errors


def export_patients(filepath):
    if openpyxl is None:
        raise ImportError("openpyxl 未安装")
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT monthly_no,medical_record_no,name,phone,age,gender,occupation,address,condition,visit_date,diagnosis,treatment,created_at FROM patients ORDER BY substr(visit_date, 1, 7) DESC, CAST(monthly_no AS INTEGER) DESC")
    rows = cur.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "患者数据"
    headers = ["月序号","病案号","姓名","电话","年龄","性别","职业","常住地址","症状","就诊日期","诊断","门诊处理","登记日期"]
    header_fill = PatternFill(start_color="5C3322", end_color="5C3322", fill_type="solid")
    header_font = Font(name="Microsoft YaHei", size=11, color="D7CCC8", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D7CCC8"),
        right=Side(style="thin", color="D7CCC8"),
        top=Side(style="thin", color="D7CCC8"),
        bottom=Side(style="thin", color="D7CCC8"),
    )
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin_border
            cell.font = Font(name="Microsoft YaHei", size=10)
    ws.auto_filter.ref = ws.dimensions
    wb.save(filepath)


def _read_headers(ws):
    """Read header row and return dict mapping cleaned header name -> column index."""
    first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if not first_row:
        return {}
    mapping = {}
    for ci, h in enumerate(first_row[0]):
        if h is not None:
            key = str(h).strip().replace('　', '').replace(' ', '')
            mapping[key] = ci
    return mapping


def import_patients(filepath):
    if openpyxl is None:
        raise ImportError("openpyxl 未安装")
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    headers = _read_headers(ws)
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    # Header name aliases -> canonical field
    FIELD_KEYS = {
        "monthly_no": ["月序号"],
        "medical_record_no": ["病案号"],
        "name": ["姓名"],
        "phone": ["电话"],
        "age": ["年龄"],
        "gender": ["性别"],
        "occupation": ["职业"],
        "address": ["常住地址", "居住地址", "地址"],
        "condition": ["症状", "病因病情"],
        "visit_date": ["就诊日期"],
        "diagnosis": ["诊断"],
        "treatment": ["门诊处理"],
    }

    def _col(field):
        for alias in FIELD_KEYS.get(field, []):
            if alias in headers:
                return headers[alias]
        return None

    def _str(row, field):
        ci = _col(field)
        if ci is None:
            return ""
        val = row[ci] if ci < len(row) else None
        return str(val).strip() if val else ""

    def _date(row, field):
        import datetime as _dt
        ci = _col(field)
        if ci is None:
            return ""
        val = row[ci] if ci < len(row) else None
        if val is None:
            return ""
        if isinstance(val, _dt.datetime):
            return val.strftime("%Y-%m-%d")
        s = str(val).strip()
        if ' ' in s:
            s = s.split(' ')[0]
        return s

    def _int(row, field):
        ci = _col(field)
        if ci is None:
            return 0
        val = row[ci] if ci < len(row) else None
        try:
            return int(val) if val else 0
        except (ValueError, TypeError):
            return 0

    name_col = _col("name")
    if name_col is None:
        return 0, ["无法识别表头：缺少'姓名'列"]

    # First pass: gather all valid rows and count per visit month
    parsed = []
    month_counts = {}
    for ri, row in enumerate(rows, 2):
        if not row:
            continue
        name = _str(row, "name")
        if not name:
            continue
        visit_date = _date(row, "visit_date")
        month_key = visit_date[:7] if visit_date else ""
        entry = {
            "ri": ri, "row": row, "name": name, "visit_date": visit_date,
            "month_key": month_key, "monthly_no": "",
        }
        parsed.append(entry)
        if month_key:
            month_counts[month_key] = month_counts.get(month_key, 0) + 1

    # Sort by visit_date so within each month, earliest = 01, latest = highest
    parsed.sort(key=lambda e: e["visit_date"])
    # Auto-assign monthly_no: within each month, count up from 01
    month_pos = {}
    for entry in parsed:
        mk = entry["month_key"]
        if mk:
            pos = month_pos.get(mk, 0) + 1
            month_pos[mk] = pos
            entry["monthly_no"] = str(pos).zfill(2)

    conn = get_conn()
    cur = conn.cursor()
    success = 0
    errors = []
    for entry in parsed:
        row = entry["row"]
        ri = entry["ri"]
        try:
            name = entry["name"]
            monthly_no = entry["monthly_no"]
            visit_date = entry["visit_date"]
            phone = _str(row, "phone")
            age = _int(row, "age")
            gender = _str(row, "gender")
            occupation = _str(row, "occupation")
            address = _str(row, "address")
            condition = _str(row, "condition")
            medical_no = _str(row, "medical_record_no")
            diagnosis = _str(row, "diagnosis")
            treatment = _str(row, "treatment")
            cur.execute(
                "INSERT INTO patients(monthly_no,medical_record_no,name,phone,age,gender,occupation,address,condition,visit_date,diagnosis,treatment,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,date('now'))",
                [monthly_no, medical_no, name, phone, age, gender, occupation, address, condition, visit_date, diagnosis, treatment]
            )
            success += 1
        except Exception as e:
            errors.append(f"第{ri}行：{e}")
    conn.commit()
    conn.close()
    return success, errors


def generate_herb_template(filepath):
    if openpyxl is None:
        raise ImportError("openpyxl 未安装")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "药材导入模板"
    headers = ["品名","编号","品名Ⅱ","厂家","采购","价格","售价(/g)","进货日期","保存时长(天)","效期","数量(kg)"]
    header_fill = PatternFill(start_color="5C3322", end_color="5C3322", fill_type="solid")
    header_font = Font(name="Microsoft YaHei", size=11, color="D7CCC8", bold=True)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # Example row
    example = ["黄芪","HL-001","黄芪","同仁堂","2026-001",58,0.85,"2026-05-01",730,"2028-04-30",25]
    for ci, v in enumerate(example, 1):
        ws.cell(row=2, column=ci, value=v).font = Font(name="Microsoft YaHei", size=10)
    ws.auto_filter.ref = ws.dimensions
    wb.save(filepath)


def generate_patient_template(filepath):
    if openpyxl is None:
        raise ImportError("openpyxl 未安装")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "患者导入模板"
    headers = ["月序号","病案号","姓名","电话","年龄","性别","职业","常住地址","症状","就诊日期","诊断","门诊处理"]
    example = ["01","000001","张三","13800001111",45,"男","工人","北京市朝阳区","风寒感冒","2026-05-01","感冒","口服中药"]
    header_fill = PatternFill(start_color="5C3322", end_color="5C3322", fill_type="solid")
    header_font = Font(name="Microsoft YaHei", size=11, color="D7CCC8", bold=True)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for ci, v in enumerate(example, 1):
        ws.cell(row=2, column=ci, value=v).font = Font(name="Microsoft YaHei", size=10)
    ws.auto_filter.ref = ws.dimensions
    wb.save(filepath)
